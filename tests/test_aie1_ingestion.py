from __future__ import annotations

from dataclasses import replace
from pathlib import Path

from openpyxl import Workbook

from modules.knowledge.src.ingestion import chunk_document, load_workbook_corpus

DATASET = Path("package/ai_workspace_dataset_vietnamese_participants.xlsm")


def test_participant_workbook_loads_all_documents_without_quarantine() -> None:
    result = load_workbook_corpus(DATASET)

    assert len(result.documents) == 40
    assert result.quarantined == ()
    assert {item.classification for item in result.documents} == {
        "Public",
        "Internal",
        "Confidential",
        "Restricted",
    }
    assert len({item.version_id for item in result.documents}) == 40


def test_chunking_preserves_security_and_version_metadata() -> None:
    document = load_workbook_corpus(DATASET).documents[0]
    chunks = chunk_document(document)

    assert len(chunks) >= 6
    assert [item.ordinal for item in chunks] == list(range(len(chunks)))
    assert len({item.chunk_id for item in chunks}) == len(chunks)
    assert all(item.document_id == document.document_id for item in chunks)
    assert all(item.version_id == document.version_id for item in chunks)
    assert all(item.classification == document.classification for item in chunks)
    assert all(item.allowed_access == document.allowed_access for item in chunks)
    assert any("3.1. Probation" in item.heading_path for item in chunks)


def test_chunk_identity_is_bound_to_immutable_version() -> None:
    document = load_workbook_corpus(DATASET).documents[0]
    changed = replace(document, version_id=document.version_id + "-next")

    assert chunk_document(document)[0].chunk_id != chunk_document(changed)[0].chunk_id


def test_invalid_classification_is_quarantined(tmp_path: Path) -> None:
    path = tmp_path / "invalid.xlsx"
    workbook = Workbook()
    documents = workbook.active
    documents.title = "Documents"
    documents.append(["document_id", "title", "department", "classification", "content_vi"])
    documents.append(["DOC-X", "Tài liệu", "Company", "Secret", "# Tài liệu\nNội dung"])
    metadata = workbook.create_sheet("Document_Metadata")
    metadata.append(
        [
            "document_id",
            "title",
            "department",
            "classification",
            "owner",
            "allowed_access",
            "last_updated",
            "tags",
            "language",
            "word_count",
        ]
    )
    metadata.append(
        ["DOC-X", "Tài liệu", "Company", "Secret", "Company", "All", "2026-01-01", "", "vi", 2]
    )
    workbook.save(path)

    result = load_workbook_corpus(path)

    assert result.documents == ()
    assert result.quarantined[0].reasons == ("UNKNOWN_CLASSIFICATION",)
