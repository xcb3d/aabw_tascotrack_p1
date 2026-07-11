"""Loader for the synthetic Vietnamese participant workbook.

The workbook is an evaluation/demo source. It does not grant production access;
its access fields are preserved as input to later policy enforcement.
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from modules.knowledge.src.ingestion.models import (
    AllowedAccess,
    Classification,
    CorpusLoadResult,
    DocumentRecord,
    QuarantineRecord,
)

DOCUMENT_HEADERS = ("document_id", "title", "department", "classification", "content_vi")
METADATA_HEADERS = (
    "document_id",
    "title",
    "department",
    "classification",
    "owner",
    "allowed_access",
    "last_updated",
    "tags",
    "language",
    "word_count",
)
CLASSIFICATIONS = frozenset({"Public", "Internal", "Confidential", "Restricted"})
ALLOWED_ACCESS_VALUES = frozenset(
    {"All", "All Employees", "Own Department", "Executive Only"}
)


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _rows_by_header(worksheet: Any, expected: tuple[str, ...]) -> Iterable[tuple[int, dict[str, Any]]]:
    header_row: int | None = None
    for row in worksheet.iter_rows():
        values = tuple(_clean(cell.value) for cell in row[: len(expected)])
        if values == expected:
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError(f"Sheet {worksheet.title!r} is missing required header {expected!r}")

    for row in worksheet.iter_rows(min_row=header_row + 1, max_col=len(expected)):
        row_values = [cell.value for cell in row]
        if not any(value is not None and _clean(value) for value in row_values):
            continue
        yield row[0].row, dict(zip(expected, row_values, strict=True))


def _metadata_index(worksheet: Any) -> tuple[dict[str, tuple[int, dict[str, Any]]], set[str]]:
    result: dict[str, tuple[int, dict[str, Any]]] = {}
    duplicates: set[str] = set()
    for row_number, row in _rows_by_header(worksheet, METADATA_HEADERS):
        document_id = _clean(row["document_id"])
        if document_id in result:
            duplicates.add(document_id)
        else:
            result[document_id] = (row_number, row)
    return result, duplicates


def _source_identity(document: dict[str, Any], metadata: dict[str, Any]) -> tuple[str, str]:
    canonical = json.dumps(
        {"document": document, "metadata": metadata},
        ensure_ascii=False,
        sort_keys=True,
        default=str,
        separators=(",", ":"),
    )
    digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    return digest, f"{_clean(document['document_id'])}-v-{digest[:16]}"


def load_workbook_corpus(path: str | Path) -> CorpusLoadResult:
    """Load and validate Documents plus its required metadata companion sheet.

    Invalid records are returned in ``quarantined`` and are never included in
    ``documents``. Workbook macros are not executed.
    """

    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(workbook_path)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=False)
    try:
        missing_sheets = {"Documents", "Document_Metadata"}.difference(workbook.sheetnames)
        if missing_sheets:
            raise ValueError(f"Workbook is missing sheets: {sorted(missing_sheets)}")

        metadata, duplicate_metadata = _metadata_index(workbook["Document_Metadata"])
        documents: list[DocumentRecord] = []
        quarantined: list[QuarantineRecord] = []
        seen_document_ids: set[str] = set()

        for row_number, row in _rows_by_header(workbook["Documents"], DOCUMENT_HEADERS):
            document_id = _clean(row["document_id"])
            reasons: list[str] = []
            meta_entry = metadata.get(document_id)
            if not document_id:
                reasons.append("MISSING_DOCUMENT_ID")
            elif document_id in seen_document_ids:
                reasons.append("DUPLICATE_DOCUMENT_ID")
            if document_id in duplicate_metadata:
                reasons.append("DUPLICATE_METADATA")
            if meta_entry is None:
                reasons.append("MISSING_METADATA")

            seen_document_ids.add(document_id)
            if meta_entry is not None:
                _, meta = meta_entry
                for field in ("title", "department", "classification"):
                    if _clean(row[field]) != _clean(meta[field]):
                        reasons.append(f"METADATA_MISMATCH_{field.upper()}")
                classification = _clean(row["classification"])
                allowed_access = _clean(meta["allowed_access"])
                if classification not in CLASSIFICATIONS:
                    reasons.append("UNKNOWN_CLASSIFICATION")
                if allowed_access not in ALLOWED_ACCESS_VALUES:
                    reasons.append("UNKNOWN_ALLOWED_ACCESS")
                if _clean(meta["language"]) != "vi":
                    reasons.append("UNSUPPORTED_LANGUAGE")
            else:
                meta = {}

            if not _clean(row["title"]):
                reasons.append("MISSING_TITLE")
            if not _clean(row["department"]):
                reasons.append("MISSING_DEPARTMENT")
            if not _clean(row["content_vi"]):
                reasons.append("MISSING_CONTENT")

            if reasons:
                quarantined.append(
                    QuarantineRecord("Documents", row_number, document_id or None, tuple(reasons))
                )
                continue

            source_sha256, version_id = _source_identity(row, meta)
            tags = tuple(tag.strip() for tag in _clean(meta["tags"]).split(",") if tag.strip())
            documents.append(
                DocumentRecord(
                    document_id=document_id,
                    version_id=version_id,
                    title=_clean(row["title"]),
                    department=_clean(row["department"]),
                    classification=_clean(row["classification"]),  # type: ignore[arg-type]
                    content_vi=_clean(row["content_vi"]),
                    owner=_clean(meta["owner"]),
                    allowed_access=_clean(meta["allowed_access"]),  # type: ignore[arg-type]
                    last_updated=_clean(meta["last_updated"]),
                    tags=tags,
                    language=_clean(meta["language"]),
                    declared_word_count=int(meta["word_count"] or 0),
                    source_sheet="Documents",
                    source_row=row_number,
                    source_sha256=source_sha256,
                )
            )

        orphan_metadata = sorted(set(metadata).difference(seen_document_ids))
        quarantined.extend(
            QuarantineRecord("Document_Metadata", metadata[item][0], item, ("ORPHAN_METADATA",))
            for item in orphan_metadata
        )
        return CorpusLoadResult(tuple(documents), tuple(quarantined))
    finally:
        workbook.close()
