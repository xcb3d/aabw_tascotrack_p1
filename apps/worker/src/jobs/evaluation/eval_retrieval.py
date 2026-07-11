"""Evaluate the AIE2 hybrid retriever on the workbook's public cases."""

from __future__ import annotations

import argparse
import asyncio
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from apps.worker.src.jobs.evaluation.search_demo import build_retriever
from modules.knowledge.src.retrieval.filters import RetrievalSubject, can_read_chunk


def _evaluation_rows(path: Path) -> tuple[dict[str, Any], ...]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    try:
        worksheet = workbook["Public_Evaluation"]
        expected = (
            "question_id",
            "category",
            "user_id",
            "user_role",
            "user_department",
            "question_vi",
            "expected_permission",
            "expected_document_id",
            "answer_type",
            "difficulty",
        )
        header = None
        for row in worksheet.iter_rows():
            values = tuple(str(cell.value).strip() if cell.value is not None else "" for cell in row)
            if values[: len(expected)] == expected:
                header = row[0].row
                break
        if header is None:
            raise ValueError("Public_Evaluation header was not found")
        result = []
        for row in worksheet.iter_rows(min_row=header + 1, max_col=len(expected), values_only=True):
            if not row[0]:
                continue
            result.append(dict(zip(expected, row, strict=True)))
        return tuple(result)
    finally:
        workbook.close()


async def evaluate(args: argparse.Namespace) -> dict[str, Any]:
    retriever = build_retriever(args)
    cases = _evaluation_rows(args.workbook)
    if args.limit is not None:
        cases = cases[: args.limit]
    results = []
    reciprocal_ranks: list[float] = []
    ndcg_values: list[float] = []
    recall_at_5: list[float] = []
    permission_correct = 0

    for number, case in enumerate(cases, 1):
        subject = RetrievalSubject(
            tenant_id=args.tenant_id,
            user_id=str(case["user_id"]),
            department=str(case["user_department"]),
            role=str(case["user_role"]),
        )
        expected_documents = tuple(
            item.strip()
            for item in str(case["expected_document_id"]).split(";")
            if item.strip()
        )
        expected_permission = str(case["expected_permission"])
        readable_targets = []
        for document_id in expected_documents:
            target_chunks = [
                chunk for chunk in retriever.index.chunks if chunk.document_id == document_id
            ]
            readable_targets.append(
                any(
                    can_read_chunk(
                        subject,
                        tenant_id=chunk.tenant_id,
                        department=chunk.department,
                        classification=chunk.classification,
                        allowed_access=chunk.allowed_access,
                    )
                    for chunk in target_chunks
                )
            )
        actual_permission = "Allow" if readable_targets and all(readable_targets) else "Deny"
        permission_correct += int(actual_permission == expected_permission)

        hits = await retriever.search(str(case["question_vi"]), subject, top_k=args.top_k)
        documents = []
        for hit in hits:
            if hit.chunk.document_id not in documents:
                documents.append(hit.chunk.document_id)
        document_ranks = {
            document_id: documents.index(document_id) + 1
            for document_id in expected_documents
            if document_id in documents
        }
        rank = (
            max(document_ranks.values())
            if len(document_ranks) == len(expected_documents)
            else None
        )
        if expected_permission == "Allow":
            reciprocal_ranks.append(0.0 if rank is None else 1 / rank)
            recall_at_5.append(
                sum(1 for value in document_ranks.values() if value <= 5)
                / len(expected_documents)
            )
            dcg = sum(1 / math.log2(value + 1) for value in document_ranks.values() if value <= 10)
            ideal = sum(1 / math.log2(value + 1) for value in range(1, len(expected_documents) + 1))
            ndcg_values.append(dcg / ideal)
        results.append(
            {
                "questionId": case["question_id"],
                "question": case["question_vi"],
                "expectedPermission": expected_permission,
                "actualPermission": actual_permission,
                "expectedDocumentIds": list(expected_documents),
                "documentRanks": document_ranks,
                "coverageRank": rank,
                "returnedDocumentIds": documents,
            }
        )
        print(f"evaluated {number}/{len(cases)}", file=sys.stderr)

    allow_count = len(reciprocal_ranks)
    summary = {
        "caseCount": len(cases),
        "allowCaseCount": allow_count,
        "permissionAccuracy": permission_correct / len(cases) if cases else 0.0,
        "recallAt5": sum(recall_at_5) / allow_count if allow_count else 0.0,
        "mrr": sum(reciprocal_ranks) / allow_count if allow_count else 0.0,
        "ndcgAt10": sum(ndcg_values) / allow_count if allow_count else 0.0,
        "rerankerEnabled": not args.no_reranker,
        "results": results,
    }
    if args.output is not None:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
    return summary


def parser() -> argparse.ArgumentParser:
    value = argparse.ArgumentParser(description=__doc__)
    value.add_argument("--tenant-id", default="demo-mytasco")
    value.add_argument("--top-k", type=int, default=10)
    value.add_argument("--device", choices=("auto", "cpu", "cuda"), default="auto")
    value.add_argument("--embedding-batch-size", type=int, default=16)
    value.add_argument("--reranker-batch-size", type=int, default=4)
    value.add_argument("--no-reranker", action="store_true")
    value.add_argument("--chunks", type=Path, default=Path("database/seeds/aie1/chunks.jsonl"))
    value.add_argument(
        "--documents", type=Path, default=Path("database/seeds/aie1/documents.jsonl")
    )
    value.add_argument(
        "--embeddings", type=Path, default=Path("database/seeds/aie1/embeddings.jsonl")
    )
    value.add_argument(
        "--embedding-config",
        type=Path,
        default=Path("config/models/embedding-qwen3-0.6b.json"),
    )
    value.add_argument(
        "--reranker-config",
        type=Path,
        default=Path("config/models/reranker-qwen3-0.6b.json"),
    )
    value.add_argument(
        "--workbook",
        type=Path,
        default=Path("package/ai_workspace_dataset_vietnamese_participants.xlsm"),
    )
    value.add_argument("--limit", type=int)
    value.add_argument(
        "--output",
        type=Path,
        default=Path("evaluation/artifacts/aie2-public-retrieval.json"),
    )
    return value


def main() -> None:
    args = parser().parse_args()
    try:
        summary = asyncio.run(evaluate(args))
    except Exception as exc:
        print(f"evaluation failed: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
    compact = {key: value for key, value in summary.items() if key != "results"}
    print(json.dumps(compact, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
